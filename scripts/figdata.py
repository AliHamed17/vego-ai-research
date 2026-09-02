"""Recompute the plotted values from the delivered dataset."""
from __future__ import annotations
import collections, glob, json, os

VERDICTS = {"full", "partially", "parially", "wrong", "not sure it is a uc"}


def stage3_overturn_by_verdict(root):
    import openpyxl
    counts = collections.Counter()
    for path in sorted(glob.glob(os.path.join(root, "System", "analysis", "scores_*.xlsx"))):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            if ws.title != "compliance_vectors":
                continue
            rows = list(ws.iter_rows(values_only=True))
            hdr = [str(h) for h in rows[0]]
            sc = next(i for i, h in enumerate(hdr) if h.lower().startswith("score"))
            vi = next(i for i, h in enumerate(hdr) if "compl" in h.lower())
            for r in rows[1:]:
                if r[sc] in (0, 1):
                    counts[(str(r[vi]), r[sc])] += 1
        wb.close()
    out = {}
    for v in {a for a, _ in counts}:
        kept, over = counts[(v, 1)], counts[(v, 0)]
        out[v] = {"n": kept + over, "overturned": over, "rate": over / (kept + over)}
    return out


def stage2_by_run_support_and_certainty(root):
    import openpyxl
    support = collections.Counter()
    certainty = {"accepted in full": [], "not accepted in full": []}
    for path in sorted(glob.glob(os.path.join(root, "System", "analysis", "guideline_clusters_*.xlsx"))):
        setting = os.path.basename(path)[len("guideline_clusters_"):-len(".xlsx")]
        with open(os.path.join(root, "System", "eval_output", setting, "agentB_result.json"), encoding="utf-8") as fh:
            runs = json.load(fh)["run_outputs"]
        cert = [{str(g.get("id")): g.get("mapping_certainty") for g in run.get("reference_guidelines", [])} for run in runs]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h).replace("\n", " ") for h in rows[0]]
        rid = [hdr.index(f"Run {i} ID") for i in (1, 2, 3)]
        st = hdr.index("Status")
        for r in rows[1:]:
            raw = str(r[st]).strip().lower().replace("parially", "partially") if r[st] not in (None, "") else None
            if raw not in VERDICTS:
                continue
            present = [i for i in range(3) if r[rid[i]] not in (None, "", "null")]
            support[(len(present), raw == "full")] += 1
            vals = [cert[i].get(str(r[rid[i]]).strip()) for i in present]
            vals = [v for v in vals if isinstance(v, (int, float))]
            if vals:
                certainty["accepted in full" if raw == "full" else "not accepted in full"].append(min(vals))
        wb.close()
    by_support = {}
    for k in (0, 1, 2, 3):
        full, notfull = support[(k, True)], support[(k, False)]
        if full + notfull:
            by_support[k] = {"n": full + notfull, "rejected": notfull, "rate": notfull / (full + notfull)}
    return by_support, certainty
