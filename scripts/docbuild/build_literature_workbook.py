#!/usr/bin/env python3
"""Build the PhD research literature workbook from verified research results.

Input is the JSON emitted by the literature-research workflow: one record per
source with independently-checked metadata and a verification verdict. Sources
found by more than one researcher are merged, never listed twice.

Nothing here invents a source. Every row carries the verdict it was given, and
rows that could not be fully verified are surfaced on their own sheet rather
than being quietly mixed in with confirmed ones.

Usage: build_literature_workbook.py <results.json> <out.xlsx>
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC, OUT = sys.argv[1], sys.argv[2]

NAVY = "1B2A4A"
BLUE = "2F5AA8"
PANEL = "F5F7FB"
GREEN_T, GREEN_BG = "1E8E5A", "E6F4EC"
AMBER_T, AMBER_BG = "B8860B", "FCF3DC"
RED_T, RED_BG = "B3261E", "FBE7E5"

RQ_TITLES = {
    "RQ1": "RQ1 - Selective Intervention",
    "RQ2": "RQ2 - Governed Knowledge Reuse",
    "RQ3": "RQ3 - Evaluation and Transfer",
    "methodology": "Methodology - Design Science",
    "foundation": "Foundation - VEGO-AI Related Work",
    "general": "General and Governance",
}

RQ_QUESTION = {
    "RQ1": "When and how, in variability exploration scenarios, should an agentic assessment system "
           "request human judgment so that important uncertainties are addressed without unnecessary expert burden?",
    "RQ2": "How should expert judgment - including the system's core reasoning - be represented, validated, "
           "reconciled, and stored so it can be reused transparently without unsafe generalization or loss of human authority?",
    "RQ3": "How can expert judgment be reused and transferred across different guideline-operationalization "
           "contexts without unsafe generalization or loss of human authority?",
    "methodology": "Sources that frame the research method itself: design-science research, artifact evaluation, "
                   "and empirical study design.",
    "foundation": "The related work of the published VEGO-AI paper - the framework this doctorate builds on.",
    "general": "Cross-cutting governance, guidelines, and positioning sources.",
}

VERDICT_STYLE = {
    "VERIFIED_ONLINE": (GREEN_T, GREEN_BG, "Verified"),
    "PARTIALLY_VERIFIED": (AMBER_T, AMBER_BG, "Partial"),
    "COULD_NOT_VERIFY": (RED_T, RED_BG, "UNVERIFIED"),
}


def norm(t: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (t or "").lower()).strip()


def merge(records: list[dict]) -> list[dict]:
    """Merge records describing the same work; keep the richest, union the insight."""
    by_title: OrderedDict[str, dict] = OrderedDict()
    for r in records:
        k = norm(r.get("title", ""))
        if not k:
            continue
        if k not in by_title:
            by_title[k] = dict(r)
            by_title[k]["_found_by"] = [r.get("agent", "?")]
            continue
        cur = by_title[k]
        cur["_found_by"].append(r.get("agent", "?"))
        # prefer the stronger verification verdict
        order = ["COULD_NOT_VERIFY", "PARTIALLY_VERIFIED", "VERIFIED_ONLINE"]
        if order.index(r.get("verification", "COULD_NOT_VERIFY")) > order.index(
            cur.get("verification", "COULD_NOT_VERIFY")
        ):
            cur["verification"] = r["verification"]
            cur["verification_note"] = r.get("verification_note", cur.get("verification_note", ""))
        # keep the longest text for the substantive fields
        for f in ("authors", "venue", "doi_or_url", "what_it_establishes",
                  "relevance_to_this_phd", "citation_use", "verification_note"):
            if len(str(r.get(f, ""))) > len(str(cur.get(f, ""))):
                cur[f] = r[f]
    return list(by_title.values())


def head(ws, cols):
    ws.append([c[0] for c in cols])
    for i, (title, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF", size=10.5)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def add_rows(ws, rows, ncols, verdict_col=None):
    for n, values in enumerate(rows):
        ws.append(values)
        r = ws.max_row
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9.5)
            if n % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=PANEL)
        if verdict_col:
            v = ws.cell(row=r, column=verdict_col)
            t, bg, label = VERDICT_STYLE.get(str(v.value), (AMBER_T, AMBER_BG, str(v.value)))
            v.value = label
            v.font = Font(size=9.5, bold=True, color=t)
            v.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[r].height = 62


def main() -> None:
    data = json.load(open(SRC, encoding="utf-8"))
    sources = merge(data["sources"])
    people = data.get("key_researchers", [])

    # merge duplicate researcher entries too
    seen: OrderedDict[str, dict] = OrderedDict()
    for p in people:
        k = norm(p.get("name", ""))
        if k and k not in seen:
            seen[k] = p
        elif k and len(str(p.get("why_relevant", ""))) > len(str(seen[k].get("why_relevant", ""))):
            seen[k] = p
    people = list(seen.values())

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- READ ME ----------------
    ws = wb.create_sheet("READ ME")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 116
    t = ws.cell(row=1, column=1, value="VEGO-AI PhD - Research Literature Workbook")
    t.font = Font(bold=True, size=16, color=NAVY)
    ws.append([])
    counts = Counter(s["verification"] for s in sources)
    rq_counts = Counter(s.get("rq_tag", "general") for s in sources)
    info = [
        ("Purpose", "The per-research-question literature base for this doctorate. One row per source, tagged to the "
                    "research question it serves, with an explicit statement of what it establishes and how it relates to this "
                    "specific research - not a generic summary."),
        ("Built", "2026-08-12. Sources come from two places: the complete reference list of the published VEGO-AI paper "
                  "(Reinhartz-Berger, Bragilovski and Sturm, MODELS '26, DOI 10.1145/3822455.3830312), and targeted "
                  "research into each research question and the design-science methodology."),
        ("Verification method", "Every source was checked against an independent online record (publisher page, DOI, dblp, "
                                "arXiv or an official repository). Verdicts: Verified = record found and metadata matched; "
                                "Partial = found but some metadata differs or is unclear, see the note; UNVERIFIED = existence "
                                "could not be confirmed."),
        ("Verification result", f"{len(sources)} unique sources after merging duplicates: "
                                f"{counts.get('VERIFIED_ONLINE', 0)} verified, "
                                f"{counts.get('PARTIALLY_VERIFIED', 0)} partial, "
                                f"{counts.get('COULD_NOT_VERIFY', 0)} unverified."),
        ("No fabrication rule", "No source, DOI, author or venue in this workbook was written from memory. Anything that "
                                "could not be confirmed online is listed on the 'Needs Checking' sheet rather than being mixed "
                                "in with confirmed sources. Check that sheet before citing anything from it."),
        ("Sheets", "All Sources = the full merged list with filters. Then one sheet per research question, plus Methodology "
                   "and Foundation. Key Researchers = the people and groups whose work matters here. Coverage Analysis = "
                   "what each question now has and what is still thin. Needs Checking = every row not fully verified."),
        ("Coverage", " · ".join(f"{RQ_TITLES.get(k, k)}: {v}" for k, v in sorted(rq_counts.items()))),
        ("Claim rule", "This workbook is an inventory and analysis. The frozen database searches (QL-01 to QL-05) have NOT "
                       "been executed, so nothing here establishes review completeness or novelty. It shows what has been "
                       "found and read, not that nothing else exists."),
        ("How to use it", "Start from the research-question sheet you are writing. The 'How it relates to this research' "
                          "column is written to be usable directly as the sentence that introduces the citation."),
    ]
    for k, v in info:
        ws.append([k, v])
        r = ws.max_row
        ws.cell(row=r, column=1).font = Font(bold=True, color=NAVY, size=10.5)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(34, 15 * (len(v) // 105 + 1))

    # ---------------- All Sources ----------------
    cols = [("RQ", 12), ("Authors", 34), ("Year", 7), ("Title", 46), ("Venue", 34),
            ("DOI / URL", 34), ("Checked", 12), ("What it establishes", 60),
            ("How it relates to this research", 70), ("How to cite it here", 40)]
    ws = wb.create_sheet("All Sources")
    head(ws, cols)
    ordering = {"RQ1": 0, "RQ2": 1, "RQ3": 2, "methodology": 3, "foundation": 4, "general": 5}
    ordered = sorted(sources, key=lambda s: (ordering.get(s.get("rq_tag", "general"), 9),
                                             -int(re.sub(r"\D", "", s.get("year", "0") or "0") or 0)))
    rows = [[s.get("rq_tag", ""), s.get("authors", ""), s.get("year", ""), s.get("title", ""),
             s.get("venue", ""), s.get("doi_or_url", ""), s.get("verification", ""),
             s.get("what_it_establishes", ""), s.get("relevance_to_this_phd", ""),
             s.get("citation_use", "")] for s in ordered]
    add_rows(ws, rows, len(cols), verdict_col=7)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # ---------------- Per research question ----------------
    for tag in ["RQ1", "RQ2", "RQ3", "methodology", "foundation"]:
        subset = [s for s in ordered if s.get("rq_tag") == tag]
        if not subset:
            continue
        ws = wb.create_sheet(RQ_TITLES[tag][:31])
        ws.column_dimensions["A"].width = 116
        q = ws.cell(row=1, column=1, value=RQ_TITLES[tag])
        q.font = Font(bold=True, size=13, color=NAVY)
        ws.append([RQ_QUESTION[tag]])
        ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="5B6472")
        ws.row_dimensions[2].height = 42
        ws.append([])
        sub = [("Authors", 34), ("Year", 7), ("Title", 46), ("Venue", 32), ("DOI / URL", 32),
               ("Checked", 12), ("What it establishes", 58), ("How it relates to this research", 72)]
        ws.append([c[0] for c in sub])
        hr = ws.max_row
        for i, (title, width) in enumerate(sub, start=1):
            c = ws.cell(row=hr, column=i)
            c.font = Font(bold=True, color="FFFFFF", size=10.5)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[hr].height = 26
        ws.freeze_panes = f"A{hr + 1}"
        rows = [[s.get("authors", ""), s.get("year", ""), s.get("title", ""), s.get("venue", ""),
                 s.get("doi_or_url", ""), s.get("verification", ""), s.get("what_it_establishes", ""),
                 s.get("relevance_to_this_phd", "")] for s in subset]
        add_rows(ws, rows, len(sub), verdict_col=6)

    # ---------------- Key researchers ----------------
    ws = wb.create_sheet("Key Researchers")
    cols = [("Name", 30), ("Affiliation", 40), ("Checked", 12), ("Why they matter to this research", 100)]
    head(ws, cols)
    add_rows(ws, [[p.get("name", ""), p.get("affiliation", ""), p.get("verification", ""),
                   p.get("why_relevant", "")] for p in people], len(cols), verdict_col=3)
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    # ---------------- Coverage analysis ----------------
    ws = wb.create_sheet("Coverage Analysis")
    cols = [("Research question", 30), ("Sources", 10), ("Verified", 10), ("What it now covers", 78)]
    head(ws, cols)
    cov = []
    for tag in ["RQ1", "RQ2", "RQ3", "methodology", "foundation", "general"]:
        subset = [s for s in sources if s.get("rq_tag") == tag]
        if not subset:
            continue
        v = sum(1 for s in subset if s["verification"] == "VERIFIED_ONLINE")
        themes = Counter()
        for s in subset:
            for word in re.findall(r"\b(defer|calibrat|uncertain|active learning|alert|provenance|RLHF|"
                                   r"annotat|memory|retrieval|transfer|leakage|validity|design science|"
                                   r"evaluat|variability|grading|agent)\w*",
                                   (s.get("what_it_establishes", "") + " " + s.get("title", "")), re.I):
                themes[word.lower()] += 1
        top = ", ".join(f"{k} ({n})" for k, n in themes.most_common(8)) or "-"
        cov.append([RQ_TITLES.get(tag, tag), len(subset), v, top])
    add_rows(ws, cov, len(cols))
    ws.append([])
    note = ws.max_row + 1
    ws.cell(row=note, column=1, value="Note")
    ws.cell(row=note, column=1).font = Font(bold=True, color=NAVY)
    ws.cell(row=note, column=2,
            value="Counts describe what has been found and verified, not what exists. The frozen "
                  "searches QL-01 to QL-05 remain unexecuted, so no completeness or novelty claim "
                  "follows from these numbers.")
    ws.merge_cells(start_row=note, start_column=2, end_row=note, end_column=4)
    ws.cell(row=note, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note].height = 44

    # ---------------- Needs checking ----------------
    ws = wb.create_sheet("Needs Checking")
    cols = [("Authors", 30), ("Year", 7), ("Title", 46), ("Venue", 30),
            ("Checked", 12), ("What was found / what is unclear", 92)]
    head(ws, cols)
    flagged = [s for s in ordered if s["verification"] != "VERIFIED_ONLINE"]
    add_rows(ws, [[s.get("authors", ""), s.get("year", ""), s.get("title", ""), s.get("venue", ""),
                   s.get("verification", ""), s.get("verification_note", "")] for s in flagged],
             len(cols), verdict_col=5)
    ws.append([])
    n = ws.max_row + 1
    ws.cell(row=n, column=1, value="Check every row above against the publisher record before citing it.")
    ws.cell(row=n, column=1).font = Font(bold=True, color=RED_T)
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=6)

    wb.save(OUT)
    print(f"WROTE {OUT}")
    print(f"  unique sources: {len(sources)} (from {len(data['sources'])} records; "
          f"{len(data['sources']) - len(sources)} duplicates merged)")
    print(f"  verified {counts.get('VERIFIED_ONLINE', 0)} | partial {counts.get('PARTIALLY_VERIFIED', 0)} "
          f"| unverified {counts.get('COULD_NOT_VERIFY', 0)}")
    print(f"  researchers: {len(people)} | sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
