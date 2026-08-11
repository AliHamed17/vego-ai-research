#!/usr/bin/env python3
"""Build the "Researchers Related to This Research" workbook.

Merges the verified researcher records from the literature research pass with the
tier/relevance analysis, ranks by closeness to this specific doctorate, and states
for each person which part of the work they connect to.

No affiliation or claim is written from memory: every field comes from a record that
was checked online, and anything uncertain keeps its uncertain verdict.

Usage: build_researchers_workbook.py <lit_results.json> <relevance.json|-> <out.xlsx>
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LIT, REL, OUT = sys.argv[1], sys.argv[2], sys.argv[3]

NAVY, PANEL = "1B2A4A", "F5F7FB"
T1_T, T1_BG = "B3261E", "FBE7E5"     # direct - read first
T2_T, T2_BG = "B8860B", "FCF3DC"     # adjacent method
T3_T, T3_BG = "2F5AA8", "EAF0FB"     # domain context
T4_T, T4_BG = "5B6472", "F0F1F3"     # background
GREEN_T, GREEN_BG = "1E8E5A", "E6F4EC"

TIER_STYLE = {
    "1": (T1_T, T1_BG), "2": (T2_T, T2_BG), "3": (T3_T, T3_BG), "4": (T4_T, T4_BG),
}
TIER_LABEL = {
    "1": "1 - DIRECT",
    "2": "2 - Adjacent method",
    "3": "3 - Domain context",
    "4": "4 - Background",
}


def norm(s: str) -> str:
    return re.sub(r"[^a-z]+", " ", (s or "").lower()).strip()


def tier_of(rec: dict) -> str:
    raw = str(rec.get("tier", "")) + " " + str(rec.get("why_relevant", ""))
    m = re.search(r"\b(?:tier\s*)?([1-4])\b", str(rec.get("tier", "")))
    if m:
        return m.group(1)
    if re.search(r"\bdirect\b", raw, re.I):
        return "1"
    if re.search(r"adjacent|method", raw, re.I):
        return "2"
    if re.search(r"domain|context", raw, re.I):
        return "3"
    return "4"


def head(ws, cols):
    ws.append([c[0] for c in cols])
    for i, (title, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF", size=10.5)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def main() -> None:
    lit = json.load(open(LIT, encoding="utf-8"))
    people_records = list(lit.get("key_researchers", []))
    rel_sources = []
    if REL != "-":
        rel = json.load(open(REL, encoding="utf-8"))
        people_records += list(rel.get("key_researchers", []))
        rel_sources = rel.get("sources", []) or []

    merged: OrderedDict[str, dict] = OrderedDict()
    for p in people_records:
        k = norm(p.get("name", ""))
        if not k:
            continue
        if k not in merged:
            merged[k] = dict(p)
            continue
        cur = merged[k]
        for f in ("affiliation", "why_relevant", "connects_to", "verification", "tier"):
            if len(str(p.get(f, ""))) > len(str(cur.get(f, ""))):
                cur[f] = p[f]

    people = list(merged.values())
    for p in people:
        p["_tier"] = tier_of(p)
    order = {"1": 0, "2": 1, "3": 2, "4": 3}
    people.sort(key=lambda p: (order.get(p["_tier"], 9), norm(p.get("name", ""))))

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- READ ME ----------------
    ws = wb.create_sheet("READ ME")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 118
    t = ws.cell(row=1, column=1, value="Researchers Related to This Research")
    t.font = Font(bold=True, size=16, color=NAVY)
    ws.append([])
    counts = Counter(p["_tier"] for p in people)
    for k, v in [
        ("Purpose", "The people whose work is closest to this doctorate: who to read first, who to cite, "
                    "whose group to follow, and who is most likely to review or examine this work."),
        ("How it is ranked", "Tier 1 DIRECT - works on the exact intersection of human/expert judgment and "
                             "AI-assisted modelling or assessment; closest prior art, read these first. "
                             "Tier 2 ADJACENT METHOD - supplies a mechanism this thesis will use or adapt. "
                             "Tier 3 DOMAIN CONTEXT - automated model assessment, variability, conceptual "
                             "modelling. Tier 4 BACKGROUND - broader AI, HCI and governance framing."),
        ("Counts", " · ".join(f"Tier {k}: {v}" for k, v in sorted(counts.items()))),
        ("Connection column", "States which part of THIS work each person connects to: one of the four VEGO-AI "
                              "agents, the Cheers/ParkWise UML corpus, the substantial-versus-occasional "
                              "distinction, the human-judgment layer, or the conditional medical extension."),
        ("Verification", "Affiliations were checked against an online record where possible. Anything not fully "
                         "confirmed keeps a weaker verdict - check those before citing an affiliation in writing."),
        ("Honest limit", "This is who has been found and verified, not a claim that no one else is relevant. "
                         "The frozen database searches QL-01 to QL-05 have not been run."),
    ]:
        ws.append([k, v])
        r = ws.max_row
        ws.cell(row=r, column=1).font = Font(bold=True, color=NAVY, size=10.5)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(32, 15 * (len(v) // 108 + 1))

    # ---------------- Researchers ----------------
    cols = [("Tier", 18), ("Name", 26), ("Affiliation", 38), ("Checked", 11),
            ("Why they matter to this research", 82), ("Connects to", 44)]
    ws = wb.create_sheet("Researchers")
    head(ws, cols)
    for n, p in enumerate(people):
        ws.append([TIER_LABEL.get(p["_tier"], p["_tier"]), p.get("name", ""), p.get("affiliation", ""),
                   p.get("verification", ""), p.get("why_relevant", ""), p.get("connects_to", "")])
        r = ws.max_row
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9.5)
            if n % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=PANEL)
        tc, tbg = TIER_STYLE.get(p["_tier"], (T4_T, T4_BG))
        tcell = ws.cell(row=r, column=1)
        tcell.font = Font(size=9.5, bold=True, color=tc)
        tcell.fill = PatternFill("solid", fgColor=tbg)
        vc = ws.cell(row=r, column=4)
        if str(vc.value) == "VERIFIED_ONLINE":
            vc.value = "Verified"
            vc.font = Font(size=9.5, bold=True, color=GREEN_T)
            vc.fill = PatternFill("solid", fgColor=GREEN_BG)
        elif vc.value:
            vc.value = "Check"
            vc.font = Font(size=9.5, bold=True, color=T2_T)
            vc.fill = PatternFill("solid", fgColor=T2_BG)
        ws.row_dimensions[r].height = 58
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # ---------------- Read first ----------------
    if rel_sources:
        cols = [("Authors", 32), ("Year", 8), ("Title", 58), ("Venue", 34), ("Why read this first", 76)]
        ws = wb.create_sheet("Read First")
        head(ws, cols)
        for n, s in enumerate(rel_sources):
            ws.append([s.get("authors", ""), s.get("year", ""), s.get("title", ""),
                       s.get("venue", ""), s.get("why_read_first", "")])
            r = ws.max_row
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9.5)
                if n % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=PANEL)
            ws.row_dimensions[r].height = 50

    # ---------------- Groups ----------------
    ws = wb.create_sheet("Groups and Venues")
    cols = [("Group or venue", 44), ("People", 60), ("Why it matters here", 74)]
    head(ws, cols)
    by_aff: OrderedDict[str, list] = OrderedDict()
    for p in people:
        aff = (p.get("affiliation") or "Unstated").split(";")[0].strip()
        by_aff.setdefault(aff, []).append(p.get("name", ""))
    rows = [[aff, ", ".join(names), ""] for aff, names in
            sorted(by_aff.items(), key=lambda kv: -len(kv[1])) if len(names) > 1]
    for n, row in enumerate(rows):
        ws.append(row)
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9.5)
            if n % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=PANEL)
        ws.row_dimensions[r].height = 42

    wb.save(OUT)
    print(f"WROTE {OUT}")
    print(f"  researchers: {len(people)}  tiers: {dict(sorted(counts.items()))}")
    print(f"  read-first works: {len(rel_sources)} | sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
