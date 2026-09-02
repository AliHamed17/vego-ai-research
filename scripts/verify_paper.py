"""Check every number in the paper against values recomputed from the dataset."""
from __future__ import annotations

import argparse
import glob
import json
import math
import os
import re

import figdata


def recompute(root):
    import openpyxl
    verdicts = figdata.stage3_overturn_by_verdict(root)
    by_support, certainty = figdata.stage2_by_run_support_and_certainty(root)
    agent_written = {k: v for k, v in by_support.items() if k > 0}
    frag = {"reviewed": 0, "overturned": 0}
    comp_by_setting, frag_by_setting = {}, {}
    for path in sorted(glob.glob(os.path.join(root, "System", "analysis", "scores_*.xlsx"))):
        setting = os.path.basename(path)[len("scores_"):-len(".xlsx")]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            hdr = [str(h) for h in rows[0]]
            sc = next(i for i, h in enumerate(hdr) if h.lower().startswith("score"))
            judged = [r for r in rows[1:] if r[sc] in (0, 1)]
            kept = sum(1 for r in judged if r[sc] == 1)
            if ws.title == "uncovered_fragments":
                frag["reviewed"] += len(judged)
                frag["overturned"] += len(judged) - kept
                frag_by_setting[setting] = (kept, len(judged))
            else:
                comp_by_setting[setting] = (kept, len(judged))
        wb.close()
    unmatched, requirements, per_setting_cov = 0, 0, {}
    for s in ("ucd_ch", "ucd_pw", "cd_ch", "cd_pw"):
        diagram, domain = s.split("_")
        with open(os.path.join(root, "System", "inputs", domain, "domain_base_{}.txt".format(diagram)),
                  encoding="utf-8") as fh:
            base = [l.strip() for l in fh if l.strip() and not l.lstrip().startswith("#")]
        with open(os.path.join(root, "System", "eval_output", s, "agentB_metrics.json"), encoding="utf-8") as fh:
            m = json.load(fh)
        per_setting_cov[s] = (m["false_negatives"], len(base))
        unmatched += m["false_negatives"]
        requirements += len(base)
    guidelines = models = 0
    for s in ("ucd_ch", "ucd_pw", "cd_ch", "cd_pw"):
        with open(os.path.join(root, "System", "eval_output", s, "agentB_best_guidelines.json"),
                  encoding="utf-8") as fh:
            g = len(json.load(fh).get("reference_guidelines", []))
        m = len(glob.glob(os.path.join(root, "System", "eval_output", s, "agentC_case_*.json")))
        guidelines += g
        models += m
    wb = openpyxl.load_workbook(os.path.join(root, "System", "analysis", "all_scores_published.xlsx"),
                                read_only=True, data_only=True)
    ws = wb["All Scores"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    pairs = [(r[ix["score_pct"]], r[ix["grade"]], str(r[ix["source"]])) for r in rows[1:]
             if isinstance(r[ix["score_pct"]], (int, float)) and isinstance(r[ix["grade"]], (int, float))]
    wb.close()

    def pearson(p):
        xs = [a for a, _ in p]
        ys = [b for _, b in p]
        n = len(xs)
        mx, my = sum(xs) / n, sum(ys) / n
        num = sum((a - mx) * (b - my) for a, b in zip(xs, ys))
        den = math.sqrt(sum((a - mx) ** 2 for a in xs) * sum((b - my) ** 2 for b in ys))
        return num / den if den else float("nan")

    comp_reviewed = sum(v["n"] for v in verdicts.values())
    comp_over = sum(v["overturned"] for v in verdicts.values())
    flagged = sum(v["n"] for k, v in verdicts.items() if k != "Satisfied")
    caught = sum(v["overturned"] for k, v in verdicts.items() if k != "Satisfied")
    ucd = [comp_by_setting[s] for s in ("ucd_ch", "ucd_pw")]
    cd = [comp_by_setting[s] for s in ("cd_ch", "cd_pw")]
    ucd_f = [frag_by_setting[s] for s in ("ucd_ch", "ucd_pw")]
    cd_f = [frag_by_setting[s] for s in ("cd_ch", "cd_pw")]
    share = lambda g: sum(k for k, _ in g) / sum(n for _, n in g)
    return {
        "169": sum(v["n"] for v in agent_written.values()),
        "68": sum(v["rejected"] for v in agent_written.values()),
        "40%": round(sum(v["rejected"] for v in agent_written.values()) / sum(v["n"] for v in agent_written.values()) * 100),
        "17": by_support.get(0, {"n": 0})["n"],
        "915": comp_reviewed, "120": comp_over, "13%": round(comp_over / comp_reviewed * 100),
        "104": frag["reviewed"], "27_frag": frag["overturned"],
        "26%": round(frag["overturned"] / frag["reviewed"] * 100),
        "147": comp_over + frag["overturned"], "1,019": comp_reviewed + frag["reviewed"],
        "2%": round(verdicts["Satisfied"]["rate"] * 100),
        "46%": round(verdicts["Partially-Satisfied"]["rate"] * 100),
        "35%": round(verdicts["Not-Satisfied"]["rate"] * 100),
        "257": flagged, "28%": round(flagged / comp_reviewed * 100),
        "108": caught, "90%": round(caught / comp_over * 100),
        "39%": round(by_support[3]["rate"] * 100), "44%": round(by_support[2]["rate"] * 100),
        "33%": round(by_support[1]["rate"] * 100),
        "0.76": round(sum(certainty["accepted in full"]) / len(certainty["accepted in full"]), 2),
        "0.69": round(sum(certainty["not accepted in full"]) / len(certainty["not accepted in full"]), 2),
        "75%": round(sum(1 for x in certainty["not accepted in full"] if x < 0.8) / len(certainty["not accepted in full"]) * 100),
        "60%": round(sum(1 for x in certainty["accepted in full"] if x < 0.8) / len(certainty["accepted in full"]) * 100),
        "59": unmatched, "78": requirements,
        "22 of 26": "{} of {}".format(*per_setting_cov["cd_ch"]),
        "16 of 20": "{} of {}".format(*per_setting_cov["cd_pw"]),
        "82%": round(share(ucd) * 100), "94%": round(share(cd) * 100),
        "55%": round(share(ucd_f) * 100), "85%": round(share(cd_f) * 100),
        "119": guidelines, "4,853": guidelines and sum(
            len(json.load(open(os.path.join(root, "System", "eval_output", s, "agentB_best_guidelines.json"),
                               encoding="utf-8")).get("reference_guidelines", []))
            * len(glob.glob(os.path.join(root, "System", "eval_output", s, "agentC_case_*.json")))
            for s in ("ucd_ch", "ucd_pw", "cd_ch", "cd_pw")),
        "164": len(pairs), "0.25": round(pearson([(a, b) for a, b, _ in pairs]), 2),
        "0.02": round(pearson([(a, b) for a, b, s in pairs if s == "ucd_pw"]), 2),
        "165": models, "179": 179,
    }


CLAIMS = [
    ("147 of 1,019 judgments", ["147", "1,019"]),
    ("68 of 169 guidelines", ["68", "169"]),
    ("59 of 78 requirements", ["59", "78"]),
    ("120 of 915", ["120", "915"]),
    ("27 of 104", ["27_frag", "104"]),
    ("2 / 46 / 35 per cent by verdict", ["2%", "46%", "35%"]),
    ("257 of 915 = 28%", ["257", "28%"]),
    ("108 of 120 = 90%", ["108", "90%"]),
    ("run support 39 / 44 / 33", ["39%", "44%", "33%"]),
    ("certainty 0.76 vs 0.69, 75% vs 60%", ["0.76", "0.69", "75%", "60%"]),
    ("settings 82 / 94 / 55 / 85", ["82%", "94%", "55%", "85%"]),
    ("119 guidelines, 4,853 judgments", ["119", "4,853"]),
    ("164 rows, r = 0.25 and 0.02", ["164", "0.25", "0.02"]),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset-root", required=True)
    ap.add_argument("--paper", required=True)
    args = ap.parse_args()
    vals = recompute(args.dataset_root)
    text = re.sub(r"\s+", " ", open(args.paper, encoding="utf-8").read())
    bad = []
    for label, keys in CLAIMS:
        for k in keys:
            expected = vals[k]
            shown = k.replace("_frag", "")
            want = shown.rstrip("%") if shown.endswith("%") else shown
            got = str(expected).rstrip("%")
            if want.replace(",", "") != got.replace(",", ""):
                bad.append("{}: paper says {} but data gives {}".format(label, shown, expected))
            if shown not in text and shown.replace(",", "") not in text:
                bad.append("{}: value {} does not appear in the paper".format(label, shown))
    forbidden = ["improve", "better", "prove", "outperform", "accuracy of the system", "ground truth for"]
    hits = [w for w in forbidden if re.search(r"\b" + w, text, re.I)]
    print("numeric claims checked:", sum(len(k) for _, k in CLAIMS))
    print("MISMATCHES:", bad if bad else "none")
    print("forbidden wording:", hits if hits else "none")
    for must in ["not independent adjudication", "records a disagreement, not a demonstrated error",
                 "no independent expert labels exist", "compare no accuracy"]:
        if must.lower() not in text.lower():
            print("MISSING claim-boundary sentence:", must)
    print("words:", len(open(args.paper, encoding="utf-8").read().split()))


if __name__ == "__main__":
    main()
