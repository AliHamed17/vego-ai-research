#!/usr/bin/env python3
"""Build the separate tracking workbook (deliverable A08-06) for 2026-08-12.

Iris asked for two documents kept apart: the Word proposal, and a separate
tracking/status document. This is the tracking half. Sheets:

  1. READ ME        - what this document is and how status is defined
  2. Aug-5 Requirements - E1..E15 with implementation state
  3. Action Items   - A08-01..A08-09 with owner and state
  4. Deliverables   - the Aug-12 package, item by item
  5. Decisions Open - decisions the supervisors must make
  6. Evidence Gates - the counts that bound every claim

Everything traces to the canonical machine-derived record; nothing is marked
approved that has not been approved.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
WARN = PatternFill("solid", fgColor="FFF2CC")
DONE = PatternFill("solid", fgColor="E2EFDA")
OPEN_ = PatternFill("solid", fgColor="FCE4D6")
BLOCK = PatternFill("solid", fgColor="F8CBAD")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

STATE_FILL = {
    "Done": DONE, "Applied": DONE, "Delivered": DONE, "Verified": DONE,
    "In progress": WARN, "Ali only": WARN, "Pending": WARN, "Draft ready": WARN,
    "Open": OPEN_, "Not started (instructed)": OPEN_, "Blocked": BLOCK, "Supervisor decision": OPEN_,
}


def header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def widths(ws, spec):
    for i, w in spec.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def table(ws, title, cols, rows, state_col=None, banner=None, colwidths=None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    start = 2
    if banner:
        ws.cell(row=2, column=1, value=banner)
        ws.cell(row=2, column=1).fill = WARN
        ws.cell(row=2, column=1).font = Font(bold=True, size=9.5)
        ws.cell(row=2, column=1).alignment = WRAP
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
        ws.row_dimensions[2].height = 28
        start = 3
    hrow = start + 1
    ws.append([])
    for _ in range(hrow - ws.max_row - 1):
        ws.append([])
    ws.append(cols)
    header(ws, ws.max_row, len(cols))
    hdr_row = ws.max_row
    for r in rows:
        ws.append(list(r))
        rr = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row=rr, column=c).alignment = WRAP
            ws.cell(row=rr, column=c).border = BORDER
        ws.cell(row=rr, column=1).font = Font(bold=True)
        if state_col:
            val = str(ws.cell(row=rr, column=state_col).value or "")
            for key, fill in STATE_FILL.items():
                if val.startswith(key):
                    ws.cell(row=rr, column=state_col).fill = fill
                    ws.cell(row=rr, column=state_col).font = Font(bold=True, size=9.5)
                    break
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(cols))}{ws.max_row}"
    if colwidths:
        widths(ws, colwidths)


def main() -> int:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("outputs/aug12/05 - Requirements and Progress Tracking.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    # 1 READ ME
    ws = wb.create_sheet("READ ME")
    ws["A1"] = "Requirements and Progress Tracking"
    ws["A1"].font = TITLE_FONT
    info = [
        ("Deliverable", "A08-06 from the 2026-08-05 call: a tracking document kept SEPARATE from the Word proposal"),
        ("For", "Supervisor meeting Wednesday 2026-08-12, 09:00 (Iris, Arnon, Ali)"),
        ("Source of truth", "docs/research/meetings/2026-08-05-supervisor-meeting.md (canonical machine-derived record, E1-E15 + A08-01..09)"),
        ("", ""),
        ("IMPORTANT", "The 2026-08-05 record is machine-transcribed with INFERRED speakers and is pending participant "
                      "confirmation. Nothing in this workbook is marked approved. 'Applied' means the correction was "
                      "implemented in the drafts; it does not mean a supervisor signed it off."),
        ("", ""),
        ("Status vocabulary", "Applied = correction implemented in the drafts | Delivered = artefact exists in this package | "
                              "In progress = started, not finished | Not started (instructed) = deliberately not begun per Iris's "
                              "instruction | Ali only = cannot be done by anyone else | Supervisor decision = needs Iris/Arnon | "
                              "Blocked = external gate closed"),
        ("", ""),
        ("Claim boundary", "EXP-005 holds 0 of 24 required independent generalization-safe expert labels; 0 of 6 medical entry "
                           "gates pass; literature searches QL-01-QL-05 are protocol-ready and NOT run. No accuracy, "
                           "generalization, clinical-performance or effort-reduction claim is made anywhere."),
    ]
    for r, (k, v) in enumerate(info, start=3):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = WRAP
        cell = ws.cell(row=r, column=2, value=v)
        cell.alignment = WRAP
        if k in {"IMPORTANT", "Claim boundary"}:
            cell.fill = WARN
    widths(ws, {1: 22, 2: 112})

    # 2 Aug-5 requirements
    reqs = [
        ("E1", "Iris", "Move from a statement of intentions to an actual thesis proposal; time is tight", "00:00:08-00:00:35",
         "In progress", "Standing pressure; governs the whole package. v0.3 proposal + full Chapter 3 delivered this cycle."),
        ("E4", "Arnon", "Draft conflates the Agentic-AI SOLUTION with the RESEARCH QUESTION; sharpen the split; cut padding", "00:11:01-00:13:07",
         "Applied", "Chapter 3 rewritten as an open-question argument. Every gap paragraph cites literature or evidence, never VEGO-AI features. §3.9 states 'the question is the contribution, the platform is the vehicle'."),
        ("E5", "Iris", "Remove 'reused' from the main RQ - reuse belongs in a sub-question", "00:13:39-00:15:09",
         "Applied", "U-RQ carries no 'reused'; reuse lives in SQ2 (capture/govern) and SQ3 (transfer). NOTE: the proposal TITLE still contains it - D-TITLE-01 raised."),
        ("E6", "Iris", "Narrow the main RQ explicitly to variability identification/classification", "00:15:10-00:17:23",
         "Applied", "U-RQ narrowed to variability exploration in guideline operationalization scenarios. DISCREPANCY FLAGGED: record says 'identification/classification', reconstructed wording says 'exploration' - A08-01 must settle it."),
        ("E7", "Iris/Arnon/Ali", "Keep 'reliable'; lean toward dropping 'auditable'/'transferable'/'end-to-end'; insert 'variability exploration scenarios' into SQ1", "00:26:29-00:29:27",
         "Applied", "Applied per the leaning; exact final text still needs the A08-01 saved-draft check."),
        ("E8", "Iris/Arnon", "Split SQ2 into (a) what the expert says/how captured vs (b) how it transfers; 'expert' vs 'human' - leaning 'expert'", "00:29:27-00:36:29",
         "Applied", "Capture-vs-transfer split is explicit in §3.6. 'Expert judgment' used in SQ2/SQ3. ASYMMETRY FLAGGED: U-RQ and SQ1 still say 'human judgment' - raised as a D-RQ-01 item rather than harmonised unilaterally."),
        ("E9", "Iris", "Tie SQ2 explicitly to 'core reasoning'; add an evaluation-criteria clause in SQ1/SQ3 style", "00:41:31-00:44:02",
         "Applied", "'Including the system's core reasoning' is in the SQ2 wording AND argued in §3.6 prose. Evaluation criteria are built into all three sub-questions."),
        ("E10", "Iris", "SQ3 wording 'looks good'; mild discomfort with 'transparently'", "00:39:20-00:41:00",
         "Applied", "SQ3 wording kept. 'Transparently' currently sits in SQ2; retained with a flagged note for the final wording pass."),
        ("E11", "Iris", "Start from the SE domain, not medical; Clalit chronic-pain vs Soroka AMD as maximally different examples", "00:37:19-00:39:20",
         "Applied", "§3.7 uses exactly this contrast to justify a classification rather than a blanket transfer claim. SQ3 stays software/modelling-first."),
        ("E12", "Arnon", "SQ3 must classify domain-specific vs broadly transferable uncertainty (actor/use-case failure = general capability gap)", "00:52:17-00:53:54",
         "Applied", "Classification is the analytic core of §3.7. The actor/use-case example is stated as a CANDIDATE hypothesis to test, not as an established fact."),
        ("E13", "Iris", "Confirmed chapter structure: Intro (later), Lit Survey, Gap & RQ, Methodology (Design Science per Penina), Preliminary Results (SE only), Plan (per RQ)", "00:44:37-00:52:17",
         "Applied", "Proposal v0.3 follows this skeleton exactly and states the true status of every chapter."),
        ("E14", "Iris/Arnon", "Plan in ~3-month semester-aligned blocks over a 3-year horizon - never month-by-month", "00:47:49-00:49:46",
         "Not started (instructed)", "Applies to the Plan chapter, which is not this cycle's deliverable. Recorded so it is not lost."),
        ("E15", "Iris", "Next steps: Chapter 3 complete; per-RQ literature spreadsheet; think (don't start) §2+§4; share Drive; Word proposal + separate tracking doc; her check-in email; live presentation Aug 12", "00:49:46-00:56:11",
         "In progress", "See the Deliverables sheet: 5 of 7 delivered by this package; Drive share and email reply are Ali-only."),
    ]
    ws = wb.create_sheet("Aug-5 Requirements")
    table(ws, "Requirements from the 2026-08-05 supervision call (E1-E15)",
          ["ID", "Raised by", "Requirement (paraphrase)", "Timestamp", "State", "How it is addressed / what remains"],
          reqs, state_col=5,
          banner="Paraphrases of machine-generated Hebrew ASR with INFERRED speakers, pending participant confirmation. "
                 "'Applied' = implemented in the drafts, NOT supervisor-approved. E2/E3/E9(A08-09) informational rows omitted.",
          colwidths={1: 7, 2: 14, 3: 62, 4: 19, 5: 22, 6: 78})

    # 3 Action items
    actions = [
        ("A08-01", "Ali only", "Finalize the RQ/SQ wording from the live-edit session - verify against the saved working draft, not the meeting record",
         "Before Aug 12", "Ali only", "BLOCKS Chapter 3 sign-off. Two specific discrepancies now surfaced for this check: 'exploration' vs 'identification/classification' (E6), and 'human' vs 'expert' judgment (E8)."),
        ("A08-02", "Ali", "Write the Gap and Research Questions chapter in full", "Aug 12", "Delivered",
         "Complete chapter delivered with the seed-corpus citations attached, construct definitions, falsifiability section and evidence-status table."),
        ("A08-03", "Ali", "Build a per-RQ literature spreadsheet with an RQ1/RQ2/RQ3/general tagging column", "Aug 12", "Delivered",
         "40 verified sources tagged; plus the per-RQ coverage-gap analysis Iris asked for beyond an inventory, and the frozen search protocol."),
        ("A08-04", "Ali", "Think about (do NOT start) section 2 (literature survey) and section 4 (research artifact per RQ)", "Ongoing", "Delivered",
         "Options note: 4 structural options for §2, 9 artifact options across SQ1-SQ3, 14 open questions. No design committed, no search executed."),
        ("A08-05", "Ali only", "Share the project Drive with Iris and Arnon", "Before Aug 12", "Ali only",
         "Requires Ali's Google account. The package folder is populated and ready to share. Also progresses standing controls R-17/A-04/Q-07."),
        ("A08-06", "Ali", "Maintain the Word proposal AND a separate tracking document", "Ongoing", "Delivered",
         "Proposal v0.3 (Word) and this tracking workbook are separate artefacts, as instructed."),
        ("A08-07", "Iris", "Send a check-in email before the Aug 12 meeting", "Before Aug 12", "Pending",
         "Iris-side. Ali replies with real status when it arrives - the Deliverables sheet is the honest answer."),
        ("A08-08", "Ali", "Present progress live at the Aug 12 meeting", "Aug 12", "Draft ready",
         "Walkthrough outline and anticipated Q&A prepared; one dry run recommended before the meeting."),
        ("A08-09", "Iris/Arnon", "Undergraduate team scoping, TA course-pipeline work, Ma'ayanei HaYeshua direction", "Ongoing", "Not an Ali action",
         "Informational only - recorded so it is not mistaken for an Ali deliverable."),
    ]
    ws = wb.create_sheet("Action Items")
    table(ws, "Action items from the 2026-08-05 call (A08-01 to A08-09)",
          ["ID", "Owner", "Action", "Timing", "State", "Status detail"], actions, state_col=5,
          banner="'Ali only' marks the items nobody else can do: they need Ali's saved draft, his Google account, or his voice.",
          colwidths={1: 9, 2: 13, 3: 60, 4: 15, 5: 16, 6: 82})

    # 4 Deliverables
    deliverables = [
        ("00", "README - Start Here", "Delivered", "Package index and reading order"),
        ("01", "Executive Brief (EN + HE)", "Delivered", "One-page summary of where the work stands"),
        ("02", "Chapter 3 - Gap and Research Questions", "Delivered", "THE deliverable for this meeting (A08-02). Full chapter, cited, with evidence-status table"),
        ("03", "Literature Review - Per Research Question (xlsx)", "Delivered", "A08-03. 40 tagged sources + per-RQ coverage gaps + frozen search protocol"),
        ("04", "Progress Presentation", "Delivered", "Live walkthrough support for A08-08"),
        ("05", "Requirements and Progress Tracking (xlsx)", "Delivered", "This workbook - A08-06 tracking half"),
        ("06", "Sections 2 and 4 - Thinking Notes", "Delivered", "A08-04. Options only; nothing started"),
        ("07", "Proposal v0.3 (Word)", "Delivered", "A08-06 proposal half; carries the D-TITLE-01 decision"),
        ("08", "Walkthrough Outline + Anticipated Q&A", "Delivered", "Meeting support; rehearsal aid"),
        ("--", "Drive shared with Iris and Arnon", "Ali only", "A08-05 - needs Ali's Google account"),
        ("--", "Reply to Iris's check-in email", "Ali only", "A08-07 response - needs Ali"),
    ]
    ws = wb.create_sheet("Deliverables")
    table(ws, "Aug-12 package contents", ["#", "Item", "State", "Purpose"], deliverables, state_col=3,
          banner="Everything marked Delivered exists in this Drive folder. Two items remain that only Ali can complete.",
          colwidths={1: 6, 2: 52, 3: 14, 4: 78})

    # 5 Open decisions
    decisions = [
        ("D-RQ-01", "Final U-RQ and SQ1-SQ3 wording", "Supervisor decision",
         "Verify against Ali's saved working draft (A08-01). Includes: 'exploration' vs 'identification/classification'; 'human' vs 'expert' judgment; whether 'transparently' stays."),
        ("D-RQ-02", "Ratify or reverse the E7 leaning (drop auditable/transferable/end-to-end from the headline)", "Supervisor decision",
         "Currently applied per the leaning. Formal sign-off outstanding."),
        ("D-TITLE-01", "Proposal title", "Supervisor decision",
         "v0.2 title still contains 'Reusable', 'Auditable', 'Transferable' - the words the RQ headline dropped. Three candidate titles proposed in the proposal document."),
        ("D-DECOMP-01", "Is the three-sub-question decomposition jointly sufficient?", "Supervisor decision",
         "None of SQ1-SQ3 takes 'reliability of the co-reasoning' as its object, though the main question names it as the outcome. Raised explicitly in Chapter 3 §3.8 rather than hidden."),
        ("D-CH2-01", "Chapter 2 structure", "Supervisor decision",
         "Four options prepared (SQ-parallel / community-parallel / lifecycle-stage / hybrid). Choice interacts with the wording sign-off."),
        ("D-ART-01", "Artifact granularity and abstraction per study", "Supervisor decision",
         "Does each study need exactly one named artifact, or is a package acceptable? Study 3's current 'Core artifact' row lists ten items, several of which are study outputs rather than artifacts."),
        ("D-SEARCH-01", "Do QL-01-QL-05 execute before or after the wording sign-off?", "Supervisor decision",
         "The queries were frozen against the provisional wording. Coverage analysis also identifies concepts no frozen query covers (e.g. learning-to-defer, LLM uncertainty calibration) - add QL-06+?"),
        ("D-PRELIM-01", "Which preliminary results may appear in the proposal, and labelled how?", "Supervisor decision",
         "Offline replay evidence exists; synthetic fixtures remain protocol-unapproved. Also: is instrument-reliability evidence admissible ahead of EXP-005 labels?"),
    ]
    ws = wb.create_sheet("Decisions Open")
    table(ws, "Decisions required from the supervisors", ["ID", "Decision", "State", "Context and options"],
          decisions, state_col=3,
          banner="These are the questions to work through on Aug 12. Each is stated so it can be answered yes/no or by choosing a named option.",
          colwidths={1: 14, 2: 56, 3: 20, 4: 92})

    # 6 Evidence gates
    gates = [
        ("EXP-005 generalization-safe expert labels", "0 of 24 supplied", "Blocked",
         ">=20 required before ANY quantitative accuracy claim; 1-19 would be pilot-only"),
        ("Medical entry gates G1-G6", "0 of 6 pass", "Blocked",
         "No medical data processing of any kind before all six pass. Internal checkpoint 2026-08-26 decides Plan A vs Plan B"),
        ("Literature searches QL-01-QL-05", "Protocol ready / NOT run", "Open",
         "No novelty or review-completeness statement is permitted until executed, screened and appraised"),
        ("Accuracy / generalization claims", "None made", "Verified",
         "Absence of such claims is itself checked by an automated evidence-consistency guard"),
        ("Clinical performance claims", "None made", "Verified", "No medical performance result exists"),
        ("RQ/SQ wording", "Provisional", "Open", "Pending A08-01 verification and logged D-RQ-01/D-RQ-02"),
        ("2026-08-05 meeting record", "Machine-derived, speakers inferred", "Open",
         "Pending participant confirmation; verbatim attribution not permitted until confirmed"),
    ]
    ws = wb.create_sheet("Evidence Gates")
    table(ws, "Evidence gates - the counts that bound every claim", ["Gate", "Current value", "State", "What it blocks"],
          gates, state_col=3,
          banner="These numbers are the honest answer to 'how far along are you'. They are unchanged by this package.",
          colwidths={1: 42, 2: 30, 3: 12, 4: 84})

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"tracking workbook: 6 sheets -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
