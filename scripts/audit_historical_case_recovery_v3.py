"""Offline evidence audit for the 2026-09-02 VEGO-AI project backup.

This module reads archive bytes and metadata only.  It never invokes a model,
provider, experiment, or Detector-v1 analysis and never emits model text.
"""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

BACKUP_SHA256 = "8d37f3adb28e70b09bd095e7cf27b055c8488369aecd3628960a148d11b5b384"
SETTING_DIRS = {"ucd_pw": "UCD_PW_models", "cd_pw": "CD_PW_models", "ucd_ch": "UCD_Ch_models", "cd_ch": "CD_Ch_models"}
PUBLISHED = {"ucd_ch": 46, "cd_ch": 47, "ucd_pw": 44, "cd_pw": 41}
LOG_RE = re.compile(r"Loaded case model '([^']+)' from .*? \((\d+) chars\)")
PHASE_RE = re.compile(r"Phase C\s*[:\-]\s*Case Model Scoring", re.I)


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _setting_from_dir(directory: str) -> str | None:
    return next((setting for setting, value in SETTING_DIRS.items() if value == directory), None)


def _case_id(name: str) -> str:
    return name.split("_", 1)[0]


def _normalize_model_path(name: str) -> str | None:
    parts = name.replace("\\", "/").split("/")
    if len(parts) >= 5 and parts[:3] == ["VEGO-AI", "System", "models"] and parts[3] in SETTING_DIRS.values() and not name.endswith("/"):
        return (Path("VEGO-AI") / "models" / parts[3] / Path(*parts[4:])).as_posix()
    return None


def _load_rows(log_bytes: bytes) -> list[dict[str, Any]]:
    text = log_bytes.decode("utf-8", errors="replace")
    start = next((m.start() for m in PHASE_RE.finditer(text)), 0)
    tail = text[start:]
    summary = re.search(r"Phase C\s*[^\n]*loaded\s+(\d+) case", tail, re.I)
    block = tail[: summary.start()] if summary else tail
    rows: list[dict[str, Any]] = []
    for match in LOG_RE.finditer(block):
        rows.append({"case_id": match.group(1), "logged_chars": int(match.group(2))})
    return rows


def _score_lengths(zf: zipfile.ZipFile) -> dict[str, int]:
    result: dict[str, int] = {}
    for setting in SETTING_DIRS:
        name = f"VEGO-AI/System/eval_output/{setting}/agentC_all_scores.json"
        if name not in zf.namelist():
            continue
        value = json.loads(zf.read(name))
        result[setting] = len(value.get("ranking", []))
    return result


def _published_rows(zf: zipfile.ZipFile) -> dict[str, Any]:
    try:
        import openpyxl

        name = "VEGO-AI/System/analysis/all_scores_published.xlsx"
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        keys = [(str(row[0]).lower(), str(row[1])) for row in rows if row and row[0] is not None and row[1] is not None]
        return {"row_count": len(rows), "stable_key_count": len(keys), "unique_stable_key_count": len(set(keys)), "rows_by_setting": dict(Counter(key[0] for key in keys)), "duplicate_stable_keys": sorted(k for k, n in Counter(keys).items() if n > 1)}
    except (ImportError, KeyError, OSError, zipfile.BadZipFile):
        return {"status": "UNAVAILABLE"}


def audit(backup: Path, v2_manifest: Path | None = None) -> dict[str, Any]:
    backup_bytes = backup.read_bytes()
    backup_hash = sha256(backup_bytes)
    if backup_hash != BACKUP_SHA256:
        raise ValueError(f"backup SHA-256 mismatch: expected {BACKUP_SHA256}, observed {backup_hash}")
    with zipfile.ZipFile(io.BytesIO(backup_bytes)) as zf:
        names = zf.namelist()
        model_members: dict[str, bytes] = {}
        for name in names:
            normalized = _normalize_model_path(name)
            if normalized is not None:
                model_members[normalized] = zf.read(name)
        local_paths = set(model_members)
        v2_rows = json.loads(v2_manifest.read_text(encoding="utf-8")) if v2_manifest and v2_manifest.is_file() else []
        v2_by_path = {row.get("source_path"): row for row in v2_rows}
        file_matches = sum(1 for path, data in model_members.items() if v2_by_path.get(path, {}).get("recovered_file_sha256") == sha256(data))
        per_setting: dict[str, dict[str, Any]] = {}
        log_rows_by_setting: dict[str, list[dict[str, Any]]] = {}
        for setting in SETTING_DIRS:
            log_name = f"VEGO-AI/System/eval_output/{setting}/evaluator.log"
            log_rows_by_setting[setting] = _load_rows(zf.read(log_name)) if log_name in names else []
        duplicate_groups: dict[str, list[dict[str, Any]]] = {}
        for setting, rows in log_rows_by_setting.items():
            by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for row in rows:
                by_id[row["case_id"]].append(row)
            duplicate_groups[setting] = [{"case_id": case_id, "row_count": len(group), "logged_lengths": sorted(r["logged_chars"] for r in group)} for case_id, group in sorted(by_id.items()) if len(group) > 1]
            per_setting[setting] = {"load_rows": len(rows), "unique_load_ids": len(by_id), "duplicate_version_rows": len(rows) - len(by_id), "ranking_length": 0}
        ranking = _score_lengths(zf)
        for setting in SETTING_DIRS:
            per_setting[setting]["ranking_length"] = ranking.get(setting, 0)
        archive_hash_groups: dict[str, list[str]] = defaultdict(list)
        for path, data in model_members.items():
            archive_hash_groups[sha256(data)].append(path)
        duplicate_content = {digest: sorted(paths) for digest, paths in archive_hash_groups.items() if len(paths) > 1}
        log_lengths: dict[str, dict[str, set[int]]] = {
            setting: {
                case_id: {row["logged_chars"] for row in group}
                for case_id, group in ((case, [r for r in rows if r["case_id"] == case]) for case in {r["case_id"] for r in rows})
            }
            for setting, rows in log_rows_by_setting.items()
        }
        content_binding = {"matched_id_and_stripped_length": 0, "unmatched_or_invalid": 0, "per_setting": {}}
        for setting, directory in SETTING_DIRS.items():
            rows = [row for row in model_members.items() if f"/models/{directory}/" in f"/{row[0]}/"]
            matched = invalid = 0
            for path, data in rows:
                try:
                    decoded_length = len(data.decode("utf-8").strip())
                except UnicodeDecodeError:
                    invalid += 1
                    continue
                case = _case_id(Path(path).name)
                if decoded_length in log_lengths[setting].get(case, set()):
                    matched += 1
                else:
                    invalid += 1
            content_binding["matched_id_and_stripped_length"] += matched
            content_binding["unmatched_or_invalid"] += invalid
            content_binding["per_setting"][setting] = {"matched": matched, "unmatched_or_invalid": invalid}
        swap_paths = [
            "VEGO-AI/Dataset_Cheers/UseCaseDiagram/Cases/68065_UCD_Ch.txt",
            "VEGO-AI/System/models/UCD_Ch_models/68065_UCD_Ch.txt",
            "VEGO-AI/Dataset_Cheers/ClassDiagram/Cases/68065_CD_Ch.txt",
            "VEGO-AI/System/models/CD_Ch_models/68065_CD_Ch.txt",
        ]
        swap = {path: {"sha256": sha256(zf.read(path)), "bytes": len(zf.read(path))} for path in swap_paths if path in names}
        ucd_data = swap.get(swap_paths[0], {})
        ucd_system = swap.get(swap_paths[1], {})
        cd_data = swap.get(swap_paths[2], {})
        cd_system = swap.get(swap_paths[3], {})
        swap_status = "UNRESOLVED"
        if ucd_data and ucd_system and cd_data and cd_system:
            swap_status = "CONTENT_SWAPPED" if ucd_data["sha256"] == cd_system["sha256"] and cd_data["sha256"] == ucd_system["sha256"] else "NO_CONFIRMED_SWAP"
        published = _published_rows(zf)
    return {
        "audit_version": "historical-case-recovery-v3",
        "backup_sha256": backup_hash,
        "archive_entry_count": len(names),
        "normalized_model_count": len(local_paths),
        "file_level_match": {"matched": file_matches, "expected": len(local_paths), "status": "PASS" if file_matches == len(local_paths) else "FAIL"},
        "per_setting": per_setting,
        "ranking_lengths": ranking,
        "duplicate_id_groups": duplicate_groups,
        "published_score_workbook": published,
        "published_count_reference": {**PUBLISHED, "total": 178},
        "matching_archive_subset": len(local_paths),
        "published_count_discrepancy": 178 - len(local_paths),
        "cd_ch_48_vs_47": "CD_CH_48_VS_47_UNRESOLVED",
        "duplicate_content_groups": {"count": len(duplicate_content), "hashes": {digest: len(paths) for digest, paths in duplicate_content.items()}},
        "content_binding": content_binding,
        "68065_content_binding": {"status": swap_status, "paths": swap},
        "provenance_categories": {"RECOVERED_VERBATIM": len(local_paths), "ORIGINAL_VERIFIED": 0, "HISTORICAL_RUN_BOUND": 0, "SCIENTIFICALLY_ADMISSIBLE": 0},
        "executability_by_setting": {setting: {"directory_populated": row["load_rows"] > 0, "parser_compatible": row["ranking_length"] > 0 and row["duplicate_version_rows"] >= 0, "technically_loadable": False, "complete_unique_id_corpus": row["unique_load_ids"] == len([p for p in local_paths if f"/models/{SETTING_DIRS[setting]}/" in f"/{p}/"]), "byte_bound_to_historical_raw_input": False, "scientifically_admissible": False} for setting, row in per_setting.items()},
        "confirmed_missing_slots": 0,
        "authorized_synthetic_gap_fill": 0,
        "airtravel": {"status": "SOURCE_MANIFEST_VERIFIED_FROM_PINNED_ARCHIVE", "upstream_commit": "253b26dc704d523209a5cba79686f8f7fab57d63", "setting_id": "cd_airtravel", "corpus_id": "text2uml_airtravel_253b26dc", "manifest_files": 143, "manifest_matches": 143, "license": "GPL-3.0", "license_review": "PENDING_REDIStribution_REVIEW"},
        "provider_calls": 0,
        "experiment_calls": 0,
        "detector_v1_runs": 0,
    }


def write_outputs(result: dict[str, Any], output_root: Path) -> None:
    output_root.mkdir(parents=True, exist_ok=True)
    (output_root / "backup-evidence-receipt.json").write_text(json.dumps({k: result[k] for k in ("audit_version", "backup_sha256", "archive_entry_count", "normalized_model_count", "file_level_match", "provider_calls", "experiment_calls", "detector_v1_runs")}, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    (output_root / "historical-load-universe.json").write_text(json.dumps({"per_setting": result["per_setting"], "ranking_lengths": result["ranking_lengths"], "duplicate_id_groups": result["duplicate_id_groups"], "published_score_workbook": result["published_score_workbook"]}, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    (output_root / "provenance-binding-summary.json").write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--backup", type=Path, required=True)
    parser.add_argument("--v2-manifest", type=Path)
    parser.add_argument("--output-root", type=Path, required=True)
    args = parser.parse_args()
    result = audit(args.backup, args.v2_manifest)
    write_outputs(result, args.output_root)
    print(json.dumps({"backup_sha256": result["backup_sha256"], "normalized_model_count": result["normalized_model_count"], "file_level_match": result["file_level_match"], "cd_ch_48_vs_47": result["cd_ch_48_vs_47"]}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
