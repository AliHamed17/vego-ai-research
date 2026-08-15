#!/usr/bin/env python3
"""Build the per-RQ literature workbook (deliverable A08-03) for the 2026-08-12 meeting.

Reads the verified seed-corpus JSON produced by the analysis pass and emits an
.xlsx with four sheets:

  1. READ ME          - protocol status banner and the claim rule (searches NOT run)
  2. Sources          - one row per verified source, tagged RQ1/RQ2/RQ3/general
  3. Coverage Gaps    - per-RQ analysis of what the seed corpus does NOT cover
  4. Search Protocol  - the frozen QL-01..QL-05 queries and their not-run status

Hard rule: every row must trace to an origin file. Nothing is invented here; the
script only formats data that was independently fabrication-checked.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
RQ_FILLS = {
    "RQ1": PatternFill("solid", fgColor="DEEBF7"),
    "RQ2": PatternFill("solid", fgColor="E2EFDA"),
    "RQ3": PatternFill("solid", fgColor="FCE4D6"),
    "general": PatternFill("solid", fgColor="EDEDED"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

NOT_RUN_BANNER = (
    "STATUS: the systematic literature searches QL-01-QL-05 are PROTOCOL READY / NOT RUN. "
    "This workbook is a VERIFIED SEED CORPUS assembled by citation-chasing one paper's reference "
    "list plus a curated resource pack. It is NOT a screened systematic review and NOT a review result."
)
CLAIM_RULE = (
    "CLAIM RULE (from the literature-search execution register): an unrun query, seed row, tool page, "
    "or preprint cannot establish novelty, effectiveness, or review completeness."
)


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def autosize(ws, widths: dict[int, int]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def sheet_readme(wb: Workbook, n_sources: int) -> None:
    ws = wb.create_sheet("READ ME")
    ws["A1"] = "Literature Review - Per Research Question"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("", ""),
        ("Deliverable", "A08-03 from the 2026-08-05 supervision call (Iris: per-RQ literature spreadsheet with an RQ tag column)"),
        ("Prepared for", "Supervisor meeting, Wednesday 2026-08-12, 09:00"),
        ("Status", "SEED CORPUS - not a systematic review"),
        ("", ""),
        ("PROTOCOL STATUS", NOT_RUN_BANNER),
        ("CLAIM RULE", CLAIM_RULE),
        ("", ""),
        ("What IS verified here",
         f"All {n_sources} sources were independently fabrication-checked against their origin files. "
         "Every author, year, venue and page range matches its origin. No source is invented."),
        ("Origins used",
         "(1) the reference list [1]-[29] of the group's own MODELS/MAS4Models 2026 submission; "
         "(2) literature/hitl-resource-pack/source-manifest.csv + bibliography.bib (8 entries); "
         "(3) two in-repo files recording a CDSS source, itself marked 'to be independently verified'."),
        ("Evidence limitation",
         "Most rows are title/metadata level only (see the Verification column). Statements about what a "
         "paper does NOT do are marked 'on title-level evidence' and must be confirmed at full text."),
        ("", ""),
        ("How to use this sheet",
         "Sources = the tagged inventory Iris asked for. Coverage Gaps = the per-RQ analysis of what the "
         "seed corpus does not cover and which query must close it. Search Protocol = the frozen queries, "
         "ready to execute after the Aug-12 direction is confirmed."),
        ("Evidence gates unchanged",
         "EXP-005 holds 0 of 24 required independent generalization-safe expert labels; 0 of 6 medical entry "
         "gates pass. No accuracy, generalization, clinical-performance or effort-reduction claim is made."),
        ("RQ wording", "PROVISIONAL pending A08-01 verification and logged D-RQ-01 / D-RQ-02 decisions."),
    ]
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = WRAP
        cell = ws.cell(row=r, column=2, value=v)
        cell.alignment = WRAP
        if k in {"PROTOCOL STATUS", "CLAIM RULE"}:
            cell.fill = WARN_FILL
            cell.font = Font(bold=True)
    autosize(ws, {1: 26, 2: 108})


def sheet_sources(wb: Workbook, sources: list[dict]) -> None:
    ws = wb.create_sheet("Sources")
    ws["A1"] = "Verified seed corpus - one row per source. " + NOT_RUN_BANNER
    ws["A1"].font = Font(bold=True, size=10)
    ws["A1"].fill = WARN_FILL
    ws["A1"].alignment = WRAP
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.row_dimensions[1].height = 30

    headers = ["#", "RQ tag", "Also", "Cite key", "Authors (year)", "Title", "Venue / type",
               "Origin (traceable)", "Rel.", "What it establishes", "What it leaves open (vs our gap)",
               "Verification"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 3, len(headers))

    order = {"RQ1": 0, "RQ2": 1, "RQ3": 2, "general": 3}
    ranked = sorted(
        sources,
        key=lambda s: (order.get(str(s.get("rq_tag", "")).strip(), 9),
                       {"high": 0, "medium": 1, "low": 2}.get(str(s.get("relevance", "")).lower(), 9),
                       str(s.get("cite_key", ""))),
    )
    for i, s in enumerate(ranked, start=1):
        tag = str(s.get("rq_tag", "")).strip()
        ws.append([
            i, tag, s.get("secondary_tags", ""), s.get("cite_key", ""), s.get("authors_year", ""),
            s.get("title", ""), s.get("venue_or_type", ""), s.get("origin", ""),
            s.get("relevance", ""), s.get("what_it_establishes", ""),
            s.get("what_it_leaves_open", ""), s.get("verification_status", ""),
        ])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
        fill = RQ_FILLS.get(tag)
        if fill:
            ws.cell(row=r, column=2).fill = fill
            ws.cell(row=r, column=2).font = Font(bold=True)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{ws.max_row}"
    autosize(ws, {1: 5, 2: 9, 3: 14, 4: 16, 5: 34, 6: 46, 7: 34, 8: 26, 9: 8, 10: 60, 11: 62, 12: 17})


def sheet_gaps(wb: Workbook, gaps: list[dict]) -> None:
    ws = wb.create_sheet("Coverage Gaps")
    ws["A1"] = ("Per-RQ coverage-gap analysis - the part Iris asked for beyond an inventory: "
                "what the seed corpus does NOT cover, and which frozen query must close it.")
    ws["A1"].font = Font(bold=True, size=10)
    ws["A1"].fill = WARN_FILL
    ws["A1"].alignment = WRAP
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 30

    headers = ["Research question", "What the seed corpus covers",
               "What it does NOT cover (the gap)", "Search priority to close it"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 3, len(headers))
    for g in gaps:
        ws.append([g.get("rq", ""), g.get("covered", ""), g.get("not_covered", ""),
                   g.get("search_priority", "")])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    autosize(ws, {1: 22, 2: 62, 3: 68, 4: 62})


def sheet_protocol(wb: Workbook) -> None:
    ws = wb.create_sheet("Search Protocol")
    ws["A1"] = ("Frozen search protocol QL-01-QL-05. Every row is deliberately UNEXECUTED. "
                "Execution begins only after the Aug-12 direction is confirmed.")
    ws["A1"].font = Font(bold=True, size=10)
    ws["A1"].fill = WARN_FILL
    ws["A1"].alignment = WRAP
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.row_dimensions[1].height = 30

    headers = ["Query ID", "Frozen concept", "Primary mapping", "Target databases", "Status"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 3, len(headers))
    rows = [
        ("QL-01", "Agentic or multi-agent AI with human oversight", "SQ1 / Study 1 - intervention architecture and authority",
         "ACM DL, IEEE Xplore, Scopus, Web of Science", "Protocol ready / NOT RUN"),
        ("QL-02", "Expert feedback, knowledge capture, memory, and reusable judgment", "SQ2 / Study 2 - governed judgment lifecycle",
         "ACM DL, IEEE Xplore, Scopus, Web of Science", "Protocol ready / NOT RUN"),
        ("QL-03", "Domain modeling, assessment, variability, and conformance", "Software/modeling baseline, Studies 1-3, Plan B transfer",
         "ACM DL, IEEE Xplore, Scopus, Web of Science", "Protocol ready / NOT RUN"),
        ("QL-04", "Intervention workload, governance, trust, and evaluation", "SQ1 and SQ3 - expert effort, usability, validity, governance",
         "ACM DL, IEEE Xplore, Scopus, Web of Science", "Protocol ready / NOT RUN"),
        ("QL-05", "Clinical guidelines, CDSS overrides, alert fatigue, healthcare process mining", "Conditional Plan A literature only",
         "PubMed", "Protocol ready / NOT RUN"),
    ]
    for row in rows:
        ws.append(list(row))
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=5).fill = WARN_FILL

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value="Window").font = Font(bold=True)
    ws.cell(row=note_row, column=2,
            value="2015-2026 primary window; older seminal work only through documented snowballing. "
                  "Five seed sources predate the window (Tselonis 2005, Auxepaules 2008, Pohl 2005, "
                  "Galster 2013, Metzger 2014) and are flagged PRE-WINDOW in the Sources sheet.").alignment = WRAP
    ws.cell(row=note_row + 1, column=1, value="Full protocol").font = Font(bold=True)
    ws.cell(row=note_row + 1, column=2,
            value="docs/research/phd-proposal/literature-search-execution-register.md "
                  "(exact Boolean strings, per-database execution matrix, required execution record)").alignment = WRAP
    autosize(ws, {1: 11, 2: 52, 3: 46, 4: 40, 5: 24})


def main() -> int:
    lit_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\ahamed\AppData\Local\Temp\claude\lit.json")
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("outputs/aug12/03 - Literature Review - Per Research Question.xlsx")
    data = json.loads(lit_path.read_text(encoding="utf-8"))
    sources = data.get("sources", [])
    gaps = data.get("coverage_gaps", [])

    wb = Workbook()
    wb.remove(wb.active)
    sheet_readme(wb, len(sources))
    sheet_sources(wb, sources)
    sheet_gaps(wb, gaps)
    sheet_protocol(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"literature workbook: {len(sources)} sources, {len(gaps)} gap analyses -> {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
